import openpyxl
import os


def last_column(ws, row, to_col=100):
    for temp_last in range(1, to_col):
        each_cell = ws.cell(row=row, column=temp_last).value
        if each_cell == "Total":
            return temp_last
    else:
        lastknownval = 41
        return lastknownval


def csv_civil(wb, ws_csv):


    ws = wb["Entry Sheet"]

    lastcol = last_column(ws, 11)

    # last_row inputs (worksheet, column index, to_range, opt. from_range(default 200))
    lastrow = 38

    date_new_format = ws["D5"].value

    for n in range(13, lastrow + 1):

        eqp = ws.cell(row=n, column=7).value
        if eqp is not None:
            for col in range(9, lastcol, 4):
                RT = ws.cell(row=n, column=col).value
                if RT is None:
                    RT = 0

                OT = ws.cell(row=n, column=col + 2).value
                if OT is None:
                    OT = 0

                if RT > 0 or OT > 0:
                    CRow = ws_csv.max_row
                    # A
                 
                    # B
                    ws_csv.cell(row=CRow + 1, column=2).value = date_new_format
                    # C
             
                    # G
                    if RT > 0:
                        ws_csv.cell(row=CRow + 1, column=3).value = RT
                    # H
                    if OT > 0:
                        ws_csv.cell(row=CRow + 1, column=4).value = OT

                
                    # T udf1: timesheet number
                    ts_num = ws.cell(row=10, column=4).value
                    ws_csv.cell(row=CRow + 1, column=5).value = ts_num
                   
                    # O equipment tag num
                    ws_csv.cell(row=CRow + 1, column=1).value = eqp
                    WO = ws.cell(row=5, column=col).value
                    CC = ws.cell(row=7, column=col).value
                    if WO is not None:
                        value_w = WO
                    elif CC is not None:
                        value_w = CC
                    else:
                        value_w = f"0000"
                    # W : WO or CC number
                    ws_csv.cell(row=CRow + 1, column=6).value = value_w
                    
                    op = str(ws.cell(row=9, column=col).value)
                    sub_op = str(ws.cell(row=9, column=col + 2).value)
                    # X : Op & Sub-op
                    if op != "None" and sub_op != "None":
                        value_x = f"{op.zfill(4)}-{sub_op.zfill(4)}"
                        ws_csv.cell(row=CRow + 1, column=7).value = value_x
                    elif op != "None":
                        value_x = f"{op.zfill(4)}"
                        ws_csv.cell(row=CRow + 1, column=7).value = value_x



wb_csv = openpyxl.Workbook()
ws_csv = wb_csv.active

# csv headers
csv_row_1 = ["Row Labels",	"Date",	"Working",	"Standby",	"ZZUDF1",	"ZZUDF4", "ZZUDF5"]
for temp in range(len(csv_row_1)):
    ws_csv.cell(row=1, column=temp + 1).value = csv_row_1[temp]

rootdir = r"C:\Users\divyal\Desktop\projects\NOVA Equip Py\oct 4 - oct 10"

for subdir, dirs, files in os.walk(rootdir):
    for file in files:
        wb = openpyxl.load_workbook(os.path.join(subdir, file), data_only=True)
        csv_civil(wb, ws_csv=ws_csv)
        
        upload_file_xlsx = os.path.join(os.getcwd(), "oct4-oct10.xlsx")
        wb_csv.save(upload_file_xlsx)
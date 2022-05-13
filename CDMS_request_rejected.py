import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook

np.datetime_as_string
from docx import Document
from docx.shared import Inches
import os
import xlrd


OvrDataWB = openpyxl.load_workbook('CDMS Labour Overide Dataset.xlsx')
OvrDataWS = OvrDataWB.get_sheet_by_name('OvrData')

AutoDataWB = openpyxl.load_workbook('CDMS Labour Auto Entry Dataset.xlsx')
AutoDataWS = AutoDataWB.get_sheet_by_name('Main')

OvrReqWB = openpyxl.load_workbook('OVR_Requested.xlsx')
OvrReqWS = OvrReqWB.get_sheet_by_name('OverrideRequests')

OvrRejWB = openpyxl.load_workbook('OVR_Rejected.xlsx')
OvrRejWS = OvrRejWB.get_sheet_by_name('OverrideRejected')

for i in range(2,OvrDataWS.max_row+1):
    if OvrDataWS.cell(row=i,column=13).value == 'Uploaded Awaiting Approval':
        name = OvrDataWS.cell(row=i,column=4).value + ", "+ OvrDataWS.cell(row=i,column=3).value
        name = name.lower()
        date = OvrDataWS.cell(row=i,column=1).value
        reason = OvrDataWS.cell(row=i,column=10).value
        rt = OvrDataWS.cell(row=i,column=5).value
        ot = OvrDataWS.cell(row=i,column=6).value
        dot = OvrDataWS.cell(row=i,column=7).value
        for j in range(2,OvrReqWS.max_row+1):
            findname = (OvrReqWS.cell(row=j,column=1).value).lower() 
            if ((name == findname) and (date == OvrReqWS.cell(row=j,column=17).value) and (reason == OvrReqWS.cell(row=j,column=14).value) and (rt == OvrReqWS.cell(row=j,column=2).value) and (ot ==OvrReqWS.cell(row=j,column=4).value) and (dot == OvrReqWS.cell(row=j,column=6).value)) :
                break
            elif j == (OvrReqWS.max_row):    
                for k in range(2,OvrRejWS.max_row + 1):
                    if (name == (OvrRejWS.cell(row=k,column=1).value).lower() and date == OvrRejWS.cell(row=k,column=17).value and reason == OvrRejWS.cell(row=k,column=14).value and rt == OvrRejWS.cell(row=k,column=2).value and ot ==OvrRejWS.cell(row=k,column=4).value and dot == OvrRejWS.cell(row=k,column=6).value) :
                        break
                    elif k == (OvrRejWS.max_row ):    
                        OvrDataWS.cell(row=i,column=13).value = 'Approved'
                        OvrDataWB.save('CDMS Labour Overide Dataset.xlsx')
                        max_row = AutoDataWS.max_row
                        
                        AutoDataWS.cell(row = max_row + 1 , column = 1).value =  OvrDataWS.cell(row=i,column=1).value.strftime('%d/%m/%Y')    
                        AutoDataWS.cell(row = max_row + 1 , column = 2).value =  OvrDataWS.cell(row=i,column=2).value 
                        AutoDataWS.cell(row = max_row + 1 , column = 3).value =  OvrDataWS.cell(row=i,column=3).value 
                        AutoDataWS.cell(row = max_row + 1 , column = 4).value =  OvrDataWS.cell(row=i,column=4).value 
                        AutoDataWS.cell(row = max_row + 1 , column = 5).value =  OvrDataWS.cell(row=i,column=5).value 
                        AutoDataWS.cell(row = max_row + 1 , column = 6).value =  OvrDataWS.cell(row=i,column=6).value  
                        AutoDataWS.cell(row = max_row + 1 , column = 7).value =  OvrDataWS.cell(row=i,column=7).value   
                        AutoDataWS.cell(row = max_row + 1 , column = 8).value =  OvrDataWS.cell(row=i,column=8).value
                        AutoDataWS.cell(row = max_row + 1 , column = 9).value =  OvrDataWS.cell(row=i,column=9).value     
                        AutoDataWS.cell(row = max_row + 1 , column = 10).value =  OvrDataWS.cell(row=i,column=12).value
                        
                        AutoDataWB.save("CDMS Labour Auto Entry Dataset.xlsx")


import glob
import time
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from CDMS_Main import login_cdms, get_report, browser
from shutil import copyfile
from openpyxl.styles import numbers
import xlwings

from SQL_Connect import sql_connection

def lastRow(idx, workbook, col=1):
    """ Find the last row in the worksheet that contains data.

    idx: Specifies the worksheet to select. Starts counting from zero.

    workbook: Specifies the workbook

    col: The column in which to look for the last cell containing data.
    """
    ws = workbook.sheets[idx]

    lwr_r_cell = ws.cells.last_cell      # lower right cell
    lwr_row = lwr_r_cell.row             # row of the lower right cell
    lwr_cell = ws.range((lwr_row, col))  # change to your specified column

    if lwr_cell.value is None:
        lwr_cell = lwr_cell.end('up')    # go up until you hit a non-empty cell

    return lwr_cell.row


def main(excel_file_path):

    wb_1 = Workbook
    wb_1 = load_workbook(filename=excel_file_path)
    ws_1 = wb_1.active

    # Agreement Sheets (4600100246000100, 4600100276000100, 4600102036000100, E6504, E6767)
    Agr_tuple = (
        "4600100246000100",
        "4600100276000100",
        "4600102036000100",
        "E6504",
        # "E6767",
    )

    loc_Agr = []
    maxR = ws_1.max_row
    ws_1.delete_rows(maxR)
    ws_1.delete_rows(1, 2)
    wb_1.save
    time.sleep(1)
    r = 0
    name_dict = {}
    name_dict_co = {}
    # Delete Empty Rows in excel (OpenpyXL)

    data_frame_main = pd.read_excel(excel_file_path, sheet_name="Sheet1", header=None)


    data_frame_main[2] = pd.to_datetime(
        arg=data_frame_main[2], errors="coerce", format="%Y/%m/%d"
    ).dt.date
    data_frame_main[3] = pd.to_datetime(
        arg=data_frame_main[3], errors="coerce", format="%Y/%m/%d"
    ).dt.date

    for index, rows in data_frame_main.iterrows():
        if rows[0] == Agr_tuple[r]:
            loc_Agr.append(index)
            r += 1
            if r == 4:
                break

    loc_Agr.append(maxR)
    arg = 1

    a = 0
    b = 1
    m = len(loc_Agr)
    # Datetime

    while arg < m:
        name_dict[arg] = pd.DataFrame(data_frame_main[loc_Agr[a] + 1 : loc_Agr[b] : 2])
        name_dict_co[arg] = pd.DataFrame(
            data_frame_main[loc_Agr[a] + 2 : loc_Agr[b] + 1 : 2]
        )
        a += 1
        b += 1
        arg += 1

    with pd.ExcelWriter(  # pylint: disable=abstract-class-instantiated
        excel_file_path, engine="openpyxl"
    ) as writer:
        writer.book = wb_1
        name_dict[1].to_excel(
            writer,
            sheet_name=Agr_tuple[0],
            startrow=2,
            startcol=0,
            header=False,
            index=False,
        )
        name_dict_co[1].to_excel(
            writer,
            sheet_name=Agr_tuple[0],
            startrow=2,
            startcol=6,
            header=False,
            index=False,
        )
        name_dict[2].to_excel(
            writer,
            sheet_name=Agr_tuple[1],
            startrow=2,
            startcol=0,
            header=False,
            index=False,
        )
        name_dict_co[2].to_excel(
            writer,
            sheet_name=Agr_tuple[1],
            startrow=2,
            startcol=6,
            header=False,
            index=False,
        )
        name_dict[3].to_excel(
            writer,
            sheet_name=Agr_tuple[2],
            startrow=2,
            startcol=0,
            header=False,
            index=False,
        )
        name_dict_co[3].to_excel(
            writer,
            sheet_name=Agr_tuple[2],
            startrow=2,
            startcol=6,
            header=False,
            index=False,
        )

        dict_ws = {}

        dict_ws[1] = writer.sheets[Agr_tuple[0]]
        dict_ws[2] = writer.sheets[Agr_tuple[1]]
        dict_ws[3] = writer.sheets[Agr_tuple[2]]
        """ dict_ws[4] = writer.sheets["Agr4"]
        dict_ws[5] = writer.sheets["Agr5"] """

        sheet_formatter(dict_ws, Agr_tuple, writer)

        writer.save()
        writer.close()
        print("Finished Working")


def sheet_formatter(sheet_name_list, Agr_tuple, ExcelWriter):

    i = 0
    for i in sheet_name_list:
        # sheet_name_list[i].merge_cells("A1:K1")
        # sheet_name_list[i]["A1"].value = str(Agr_tuple[j])
        # sheet_name_list[i]["A1"].alignment = Alignment(
        #     wrap_text=True, horizontal="center", vertical="center"
        # )
        # sheet_name_list[i]["A1"].fill = PatternFill("solid", fgColor="00FFFF00")
        # sheet_name_list[i]["A1"].font = Font(bold=True, size=12)
        # sheet_name_list[i].row_dimensions[1].height = 16
        sheet_name_list[i].insert_cols(1)
        sheet_name_list[i].delete_rows(1)
        sheet_name_list[i]["A1"].value = "PO Number (with P_Code)"
        sheet_name_list[i]["B1"].value = "PO Number"
        sheet_name_list[i]["C1"].value = "PO Name"
        sheet_name_list[i]["D1"].value = "PO Start Date"
        sheet_name_list[i]["E1"].value = "PO End Date"
        sheet_name_list[i]["F1"].value = "PO Status"
        sheet_name_list[i]["G1"].value = "PO Type"
        sheet_name_list[i]["H1"].value = "CO Number"
        sheet_name_list[i]["I1"].value = "CO Name"
        sheet_name_list[i]["J1"].value = "CO Start Date"
        sheet_name_list[i]["K1"].value = "CO End Date"
        sheet_name_list[i]["L1"].value = "CO Status"
        sheet_name_list[i].column_dimensions["A"].width = 16
        sheet_name_list[i].column_dimensions["B"].width = 20
        sheet_name_list[i].column_dimensions["C"].width = 14
        sheet_name_list[i].column_dimensions["D"].width = 12
        sheet_name_list[i].column_dimensions["E"].width = 9
        sheet_name_list[i].column_dimensions["F"].width = 8
        sheet_name_list[i].column_dimensions["G"].width = 13
        sheet_name_list[i].column_dimensions["H"].width = 20
        sheet_name_list[i].column_dimensions["I"].width = 12
        sheet_name_list[i].column_dimensions["J"].width = 12
        sheet_name_list[i].column_dimensions["K"].width = 9


        ExcelWriter.save()
        j = 2
        mr = sheet_name_list[i].max_row
        # ws.column_dimensions["B"].width = 8.5

        while j < mr:
            sheet_name_list[i][
                f"A{j}"
            ].value = f'=CONCATENATE(LEFT(B{j},10),IF(LEN(H{j})=12,CONCATENATE("P",RIGHT(H{j},3)),""))'
            sheet_name_list[i][f'A{j}'].number_format = numbers.FORMAT_TEXT
            sheet_name_list[i][f'A{j}'].value = sheet_name_list[i][f'A{j}'].value
            j += 1



#PO_CO_Agreements
def load_msp_log(filepath):

    po_co_df = pd.read_excel(filepath, sheet_name='4600102036000100', header=0, engine='openpyxl', usecols='A:L')
    print(po_co_df)
    return po_co_df

def update_msp_log(cnxn, df):
    #name = table name
    name = 'PO_CO_Agreement'
    cnxn.execute(f'DELETE FROM {name}')
    time.sleep(5)
    df.to_sql(name=name, con=cnxn, schema='dbo', if_exists='append', index=False)




if __name__ == "__main__":
    engine = sql_connection()
    file_name = 'PO_CO_By_Agreements.xlsx'
    conn1 = engine.connect()
    conn2 = engine.connect()
    result = conn1.execute("SELECT [UserName],[Password] FROM [dbo].[passwords] WHERE CONVERT(VARCHAR, App) = 'CDMS_login'")
    result = result.fetchall()
    username = result[0][0]
    pass1 = result[0][1]

    result2 = conn2.execute("SELECT [UserName],[Password] FROM [dbo].[passwords] WHERE CONVERT(VARCHAR, App) = 'CDMS_Track'")
    result2 = result2.fetchall()

    username2 = result2[0][0]
    pass2 = result2[0][1]

    if username != username2:
        print('Usernames in SQL_Database tables do not match please fix before next rerun')
        Exception("Password Error")

    login_cdms(username, pass1, pass2)
    # Go to report download page
    repo_link: str = "https://cdms.exxonmobil.com/reports.aspx"
    # Report 1 (Regular Swipe Report)
    browser.get(repo_link)
    get_report("Agreements and Purchase Orders By Company")
    browser.quit()
    list_of_files = glob.glob("C:/Users/" + "momine" + "/Downloads/*.xlsx")
    excel_file_path = max(list_of_files, key=os.path.getctime)
    if os.path.exists(file_name):
        os.remove(file_name)
    os.rename(excel_file_path, file_name)
    main(file_name)

    wb = xlwings.Book(file_name)
    sht = wb.sheets

    sht = sht[-1]
    last_row = lastRow(-1,wb, 1)
    for row in range(2,last_row+2):
        temp_val = sht.range(f'A{row}').value
        sht.range(f'A{row}').value = temp_val

    wb.save('temp_file.xlsx')
    time.sleep(2)
    wb.close()
    time.sleep(10)
    print('WB Saved!!')

    cnxn = engine.connect()
    temp_df = load_msp_log('temp_file.xlsx')
    update_msp_log(cnxn, temp_df)
    cnxn.close()
    engine.dispose()




# WS_1 = "ws_Agr_" + str(x)


"""  while r < maxR:
  if ws_1.cell(row=r, column=1).value == Agr_tuple[arg]:
        arg += 1
        if arg < 1:
            new_arg = False
        else:
            new_arg = True
        first_row`[arg - 1] = [r + 1]
    else:
        new_arg = False
    if new_arg == True and arg != 1:
        frRow = first_row[arg - 2]
        last_row = r - 1
        wb_1.create_sheet("Sheet2")
        ws_list = wb_1.sheetnames
        for i in range(frRow, last_row):
            for j in range(1, 6):
                ws_list[0].cell(row=i, column=j).value = ws_1.cell(row=i, column=j)
    r += 1
 """
""" name_dict[1].to_excel(
    excel_writer=writer,
    sheet_name="NewSheet",
    startrow=1,
    startcol=0,
    header=False,
    index=False,
) """

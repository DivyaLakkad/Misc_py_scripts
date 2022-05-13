import openpyxl
import datetime

def xl_maker(path):
    wb = openpyxl.load_workbook(path)
    ws = wb['Week 1']

    ws["B2"].value = "Divya Lakkad"
    ws["B3"].value = "352070"

    today = datetime.date.today()

    idx = (today.weekday() + 1) % 7 # MON = 0, SUN = 6 -> SUN = 0 .. SAT = 6

    col = 5

    for i in range(0, 7):
        
        x = today - datetime.timedelta(7+idx - i)

        if i == 0:
            sunday = x.strftime('%m.%d.%Y')
        
        if i == 6:
            saturday = x.strftime('%m.%d.%Y')

        final_x = x.strftime('%a \n%b %d')
        ws.cell(row=7, column=col).value = final_x

        if col == int(5) or col == int(17):
            pass
        else:
            ws.cell(row=9, column=col).value = ws.cell(row=19, column=col).value = "8"

        ws["S9"].value = ws["S19"].value = "40"

        col = col + 2

    ws["B4"].value = x.strftime('%m/%d/%Y')

    wb.save(f'Lakkad, Divya {sunday} - {saturday}.pdf')

    wb.close





if __name__ == '__main__':

    ts_template = r"C:\Users\divyal\Desktop\graham\downloads\Timesheets\Timesheet Template.xlsx"

    xl_maker(ts_template)
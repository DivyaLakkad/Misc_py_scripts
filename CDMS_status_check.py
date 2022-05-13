import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook

np.datetime_as_string
from docx import Document
from docx.shared import Inches
import os
import xlrd

# def doc_status(filename):
#     new_df = pd.read_excel(filename, engine="openpyxl")
#     i = new_df.shape[0]
#     print(new_df.dtypes)
#     new_df['Status'] = new_df['Status'].astype('string')
#     print(new_df.dtypes)
#     for x in range(i):
#             if pd.isna(new_df['Override Reason'].values[x]):
#                 print(x)
#                 #new_df['Status'].values[x].replace(np.nan,'Awaiting OVR Reason Selection', inplace=True)
#                 new_df['Status'].values[x] = 'Awaiting OVR Reason Selection'
#             elif pd.isna(new_df['Override Comment'].values[x]):
#                 print(x)
#                 new_df['Status'].values[x] = 'Awaiting OVR Comments'
#                 print("done")
#             elif pd.isna(new_df['Handler'].values[x]):
#                 print(x)
#                 new_df['Status'].values[x] = 'Awaiting Handler'
#             print(new_df['Status'].values[x])   
#     new_df.to_excel("output.xlsx")
# doc_status('CDMS Labour Overide Dataset.xlsx')
def doc_status(filename):

    book = openpyxl.load_workbook(filename)
    sheet = book.get_sheet_by_name('OvrData')
    max_row=sheet.max_row

    max_column=sheet.max_column
    for i in range(2,max_row+1):
          # get pa rticular cell value    
            cell_obj=sheet.cell(row=i,column=10)
            total = sheet.cell(row=i,column=5).value + sheet.cell(row=i,column=6).value + sheet.cell(row=i,column=7).value
            print(total)
            if sheet.cell(row=i,column=10).value is None:
                sheet.cell(row=i,column=13).value = 'Awaiting OVR Reason Selection'
            elif sheet.cell(row=i,column=11).value is None:
                sheet.cell(row=i,column=13).value = 'Awaiting OVR Comments'
            elif sheet.cell(row=i,column=12).value is None:
                sheet.cell(row=i,column=13).value = 'Awaiting Handler'
            elif total <= 0:
                sheet.cell(row=i,column=13).value = 'Awaiting Time'

    book.save('CDMS Labour Overide Dataset.xlsx')

doc_status('CDMS Labour Overide Dataset.xlsx')
